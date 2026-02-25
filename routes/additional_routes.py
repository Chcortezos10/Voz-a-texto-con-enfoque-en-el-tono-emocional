"""
Rutas adicionales para transcripción
en la nube (cloud)
-exportacion de datos
-gesition de sesiones
-transcripcion en la nube
"""

import os
import json
import re
import time
import uuid
import logging
from datetime import datetime
from typing import Optional, Dict, Any, List

from fastapi import APIRouter, File, UploadFile, Form, HTTPException, Query
from fastapi.responses import Response, JSONResponse, StreamingResponse
import io

from core.transcription_cloud import (
    get_transcription_service,
    TranscriptionProvider,
    validate_api_key
)
from core.export_manager import (
    get_export_manager,
    ExportData
)

logger = logging.getLogger(__name__)

_sessions_store: Dict[str, Dict[str, Any]] = {}

router = APIRouter()

# Reportlab (Helvetica) no soporta emoji ni caracteres fuera de Latin-1.
# Esta función elimina esos caracteres antes de pasarlos a Paragraph.
_EMOJI_RE = re.compile(
    r'[\U00010000-\U0010ffff]'       # supplementary planes (emoji, etc.)
    r'|[\u2600-\u27BF]'              # misc symbols, dingbats
    r'|[\u2300-\u23FF]'              # misc technical
    r'|[\uFE00-\uFE0F]'             # variation selectors
    r'|[\u200D\u200C\u200B\uFEFF]', # zero-width joiners / BOM
    flags=re.UNICODE
)

def _safe_pdf_text(text: str) -> str:
    """Elimina caracteres no soportados por Helvetica en reportlab."""
    cleaned = _EMOJI_RE.sub('', str(text))
    # Asegurar que sea Latin-1 compatible
    return cleaned.encode('latin-1', 'replace').decode('latin-1').strip()

#Trascripcion en la nube 

@router.post("/config/api_key",tags=["configuracion"])
async def set_api_key(provider: str = Form(...), api_key: str = Form(...)) -> Dict[str, Any]:
    """
    Configura la clave de API para el proveedor de transcripción en la nube.
    """
    if provider not in ["openai", "groq"]:
        raise HTTPException(status_code=400, detail="Proveedor no válido. Use 'openai' o 'groq'")

    validation = validate_api_key(provider, api_key)
    
    if not validation.get("valid"):
        return {
            "status": "error",
            "message": validation.get("message", "API key inválida"),
            "provider": provider
        }

    #configurar el servicio de transcripcion

    service = get_transcription_service()

    if provider == "openai":
        service.set_openai_api_key(TranscriptionProvider.OPENAI, api_key)
    elif provider == "groq":
        service.set_groq_api_key(TranscriptionProvider.GROQ, api_key)
    
    return {
        "status": "success",
        "message": "Clave de API configurada correctamente",
        "provider": provider
    }

@router.post("/config/validate_api_key",tags=["configuracion"])
async def validate_api_key_endpoint(provider: str = Form(...), api_key: str = Form(...)) -> Dict[str, Any]:
    """
    Valida la clave de API para el proveedor de transcripción en la nube.
    """
    if provider not in ["openai", "groq"]:
        raise HTTPException(status_code=400, detail="Proveedor no válido. Use 'openai' o 'groq'")
    
    result = validate_api_key(provider, api_key)
    return result

@router.get("/config/api_key",tags=["configuracion"])
async def estimate_cost(
    duration_seconds: float = Query(..., description="Duración del audio en segundos"),
    provider: str = Query("local", description="Proveedor: local, openai, groq")
) -> Dict[str, Any]:
    """
    Estima el costo de la transcripción en la nube.
    """
    provied_map = {
        "local": TranscriptionProvider.LOCAL,
        "openai": TranscriptionProvider.OPENAI,
        "groq": TranscriptionProvider.GROQ
    }

    if provider not in ["local", "openai", "groq"]:
        raise HTTPException(status_code=400, detail="Proveedor no válido. Use 'local', 'openai' o 'groq'")
    
    service = get_transcription_service()
    estimate = service.estimate_cost(duration_seconds, provied_map[provider])
    return estimate

@router.get("/config/estimate_cost", tags=["configuracion"])
async def get_providers() -> Dict[str, Any]:
    """
    Obtiene el servicio de transcripción en la nube.
    """
    service = get_transcription_service()
    return {
        "providers": [
            {
                "id": "local",
                "name": "Local (Whisper)",
                "description": "Procesamiento local. Gratis pero más lento.",
                "cost_per_minute": 0.0,
                "speed": "1x tiempo real",
                "available": True,
                "requires_key": False
            },
            {
                "id": "openai",
                "name": "OpenAI Whisper API",
                "description": "API de OpenAI. Rápido y preciso.",
                "cost_per_minute": 0.006,
                "speed": "~2x más rápido que tiempo real",
                "available": service.has_api_key(TranscriptionProvider.OPENAI),
                "requires_key": True
            },
            {
                "id": "groq",
                "name": "Groq (Ultra Rápido)",
                "description": "25x más rápido que tiempo real. Casi gratis.",
                "cost_per_minute": 0.0001,
                "speed": "~25x más rápido que tiempo real",
                "available": service.has_api_key(TranscriptionProvider.GROQ),
                "requires_key": True
            }
        ]
    }


#exportacion de datos

@router.post("/export/srt", tags=["export"])
async def export_srt_form(segments: str = Form(...),
    include_emotions: bool = Form(True),
    include_speaker: bool = Form(True),
    filename: str = Form("subtitulos")
) -> Response:
    try:
        segments_list = json.loads(segments)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Formato de segmentos inválido")

    exporter = get_export_manager()

    data_dict = {"segments":segments_list}
    content = exporter.export_srt(data_dict)
    
    return Response(content=content, media_type="text/plain",headers={"Content-Disposition": f"attachment; filename={filename}.srt"})

@router.post("/export/vtt", tags=["export"])
async def export_vtt(
    segments: str = Form(...),
    include_emotions: bool = Form(True),
    include_speaker: bool = Form(True),
    filename: str = Form("subtitulos")
) -> Response:

    """
    Exporta los segmentos de la transcripción en formato VTT.
    """

    try:
        segments_list = json.loads(segments)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Formato de segmentos inválido")

    exporter = get_export_manager()

    data_dict = {"segments":segments_list}
    content = exporter.export_vtt(data_dict)

    return Response(content=content, media_type="text/vtt",headers={"Content-Disposition": f"attachment; filename={filename}.vtt"})

@router.post("/export/csv", tags=["export"])
async def export_csv(
    segments: str = Form(...),
    include_all_emotions: bool = Form(True),
    filename: str = Form("subtitulos")
) -> Response:
    """
    Exporta los segmentos de la transcripción en formato CSV.
    """

    try:
        segments_list = json.loads(segments)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Formato de segmentos inválido")

    exporter = get_export_manager()

    data_dict = {"segments":segments_list}
    content = exporter.export_csv(data_dict)

    return Response(content=content, media_type="text/csv",headers={"Content-Disposition": f"attachment; filename={filename}.csv"})

@router.post("/export/pdf", tags=["export"])
async def export_pdf_endpoint(
    data: str = Form(...),
    filename: str = Form("reporte_analisis")
) -> Response:
    """
    Exporta reporte PDF profesional con scoring, resumen y emociones.
    El campo 'data' debe ser JSON con la respuesta completa del análisis.
    """
    try:
        analysis_data = json.loads(data)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="JSON inválido")

    try:
        pdf_bytes = _generate_professional_pdf(analysis_data)
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}.pdf"}
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="reportlab no instalado")
    except Exception as e:
        logger.error(f"Error generando PDF: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/export/pdf-general", tags=["export"])
async def export_pdf_general_endpoint(
    filename: str = Query("informe_general_calidad"),
    include_ai_interpretation: bool = Query(True)
) -> Response:
    """
    Genera un informe PDF profesional CONSOLIDADO de TODOS los audios procesados.
    Lee el historial completo y genera un reporte con:
    - Score general, KPIs por dimensión
    - Distribución de calidad y emociones
    - Top mejores/peores audios
    - Problemas comunes y recomendaciones
    - Interpretación AI con Qwen2.5 (si Ollama está disponible)
    """
    from pathlib import Path
    from dataclasses import asdict
    from core.scoring_engine import calculate_general_quality_metrics

    history_file = Path("data/analysis_history.json")

    try:
        if not history_file.exists():
            raise HTTPException(status_code=404, detail="No hay historial de análisis")

        with open(history_file, "r", encoding="utf-8") as f:
            history = json.load(f)

        if not history:
            raise HTTPException(status_code=404, detail="El historial está vacío")

        # Calcular métricas generales
        general_metrics = calculate_general_quality_metrics(history)
        metrics_dict = asdict(general_metrics)

        # Intentar obtener interpretación AI si se solicita
        ai_interpretation = None
        if include_ai_interpretation:
            try:
                from core.ollama_interpreter import generate_data_interpretation
                ai_interpretation = await generate_data_interpretation(metrics_dict)
            except Exception as e:
                logger.warning(f"No se pudo generar interpretación AI: {e}")

        # Generar PDF
        pdf_bytes = _generate_general_pdf(metrics_dict, history, ai_interpretation)

        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}.pdf"}
        )
    except HTTPException:
        raise
    except ImportError:
        raise HTTPException(status_code=500, detail="reportlab no instalado")
    except Exception as e:
        logger.error(f"Error generando PDF general: {e}")
        raise HTTPException(status_code=500, detail=str(e))


def _generate_general_pdf(
    metrics: dict,
    history: list,
    ai_interpretation: str = None
) -> bytes:
    """Genera un PDF profesional consolidado con todas las métricas generales."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
        Image, HRFlowable, KeepTogether, PageBreak
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.graphics.shapes import Drawing, Rect, String, Circle
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics.charts.piecharts import Pie

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        leftMargin=0.6*inch, rightMargin=0.6*inch,
        topMargin=0.5*inch, bottomMargin=0.5*inch
    )
    elements = []
    styles = getSampleStyleSheet()

    # ── Colores corporativos ──
    PRIMARY = colors.HexColor("#1a237e")
    ACCENT = colors.HexColor("#42a5f5")
    SUCCESS = colors.HexColor("#66bb6a")
    WARNING = colors.HexColor("#ffa726")
    DANGER = colors.HexColor("#ef5350")
    LIGHT_BG = colors.HexColor("#f5f5f5")
    DARK_BG = colors.HexColor("#263238")

    EMOTION_COLORS = {
        "feliz": colors.HexColor("#4caf50"),
        "neutral": colors.HexColor("#42a5f5"),
        "triste": colors.HexColor("#7e57c2"),
        "enojado": colors.HexColor("#ef5350"),
    }

    # ── Custom styles ──
    title_style = ParagraphStyle(
        'GenTitle', parent=styles['Title'],
        fontSize=22, textColor=PRIMARY, spaceAfter=4, alignment=TA_CENTER
    )
    subtitle_style = ParagraphStyle(
        'GenSubtitle', parent=styles['Heading2'],
        fontSize=14, textColor=ACCENT, spaceBefore=14, spaceAfter=6
    )
    section_style = ParagraphStyle(
        'GenSection', parent=styles['Heading3'],
        fontSize=12, textColor=PRIMARY, spaceBefore=10, spaceAfter=4
    )
    body_style = ParagraphStyle(
        'GenBody', parent=styles['Normal'],
        fontSize=10, leading=14
    )
    small_style = ParagraphStyle(
        'GenSmall', parent=styles['Normal'],
        fontSize=8, leading=10, textColor=colors.grey
    )
    center_style = ParagraphStyle(
        'GenCenter', parent=styles['Normal'],
        fontSize=10, alignment=TA_CENTER
    )

    date_str = datetime.now().strftime("%d/%m/%Y %H:%M")

    # ══════════════════════════════════════
    # PORTADA
    # ══════════════════════════════════════
    logo_path = None
    try:
        import glob
        logo_files = glob.glob(os.path.join("data", "company_logo.*"))
        if logo_files:
            logo_path = logo_files[0]
    except:
        pass

    if logo_path and os.path.exists(logo_path):
        try:
            img = Image(logo_path, width=2*inch, height=1*inch)
            img.hAlign = 'CENTER'
            elements.append(img)
            elements.append(Spacer(1, 10))
        except:
            pass

    elements.append(Spacer(1, 40))
    elements.append(Paragraph("INFORME GENERAL DE CALIDAD", title_style))
    elements.append(Paragraph("Análisis Consolidado de Atención al Cliente", ParagraphStyle(
        'CoverSub', parent=styles['Normal'],
        fontSize=14, textColor=ACCENT, alignment=TA_CENTER, spaceAfter=20
    )))
    elements.append(HRFlowable(width="60%", thickness=3, color=PRIMARY, spaceAfter=20))

    # Info box de portada
    total_audios = metrics.get("total_audios", 0)
    avg_score = metrics.get("avg_score", 0)
    gen_class = metrics.get("general_classification", "N/A")
    score_color = SUCCESS if avg_score >= 70 else WARNING if avg_score >= 50 else DANGER

    cover_data = [
        ["Total Audios Analizados", str(total_audios)],
        ["Score General Promedio", f"{avg_score}/100"],
        ["Clasificación General", gen_class],
        ["Fecha del Informe", date_str],
    ]
    cover_table = Table(cover_data, colWidths=[3*inch, 3*inch])
    cover_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), PRIMARY),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.white),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('BACKGROUND', (1, 1), (1, 1), score_color),
        ('TEXTCOLOR', (1, 1), (1, 1), colors.white),
        ('FONTNAME', (1, 1), (1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (1, 1), (1, 1), 14),
        ('BACKGROUND', (1, 0), (1, 0), LIGHT_BG),
        ('BACKGROUND', (1, 2), (1, 2), LIGHT_BG),
        ('BACKGROUND', (1, 3), (1, 3), LIGHT_BG),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.white),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
    ]))
    elements.append(cover_table)
    elements.append(PageBreak())

    # ══════════════════════════════════════
    # RESUMEN EJECUTIVO
    # ══════════════════════════════════════
    elements.append(Paragraph("1. RESUMEN EJECUTIVO", subtitle_style))
    elements.append(HRFlowable(width="100%", thickness=1, color=ACCENT, spaceAfter=8))

    std_dev = metrics.get("std_deviation", 0)
    min_score = metrics.get("min_score", 0)
    max_score = metrics.get("max_score", 0)

    exec_data = [
        ["Métrica", "Valor"],
        ["Total Audios", str(total_audios)],
        ["Score Promedio", f"{avg_score}/100"],
        ["Clasificación", gen_class],
        ["Desviación Estándar", str(std_dev)],
        ["Score Mínimo", str(min_score)],
        ["Score Máximo", str(max_score)],
    ]
    exec_table = Table(exec_data, colWidths=[3.5*inch, 3.5*inch])
    exec_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), PRIMARY),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BACKGROUND', (0, 1), (0, -1), LIGHT_BG),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#e0e0e0")),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
    ]))
    elements.append(exec_table)
    elements.append(Spacer(1, 12))

    # ══════════════════════════════════════
    # 2. KPIs POR DIMENSIÓN
    # ══════════════════════════════════════
    elements.append(Paragraph("2. KPIs POR DIMENSIÓN", subtitle_style))
    elements.append(HRFlowable(width="100%", thickness=1, color=ACCENT, spaceAfter=8))

    dim_details = metrics.get("dimension_details", {})
    if dim_details:
        dim_header = [["Dimensión", "Promedio", "Máximo", "Cumplimiento"]]
        for dim_key, details in dim_details.items():
            if isinstance(details, dict):
                label = details.get("label", dim_key)
                avg_val = details.get("average_score", 0)
                max_val = details.get("max_score", 0)
                pct = details.get("percentage", 0)
                dim_header.append([label, f"{avg_val}", f"{max_val}", f"{pct}%"])

        dim_table = Table(dim_header, colWidths=[2.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), PRIMARY),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#e0e0e0")),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ]
        # Color-code percentage cells
        for i, (dim_key, details) in enumerate(dim_details.items()):
            if isinstance(details, dict):
                row = i + 1
                pct = details.get("percentage", 0)
                if pct >= 70:
                    style_cmds.append(('BACKGROUND', (3, row), (3, row), SUCCESS))
                    style_cmds.append(('TEXTCOLOR', (3, row), (3, row), colors.white))
                elif pct >= 50:
                    style_cmds.append(('BACKGROUND', (3, row), (3, row), WARNING))
                    style_cmds.append(('TEXTCOLOR', (3, row), (3, row), colors.white))
                else:
                    style_cmds.append(('BACKGROUND', (3, row), (3, row), DANGER))
                    style_cmds.append(('TEXTCOLOR', (3, row), (3, row), colors.white))

        dim_table.setStyle(TableStyle(style_cmds))
        elements.append(dim_table)

    # Detalles de protocolo
    protocol_detail = dim_details.get("protocol", {})
    if protocol_detail:
        elements.append(Spacer(1, 8))
        elements.append(Paragraph("Cumplimiento de Protocolo:", section_style))
        proto_data = [
            ["Aspecto", "Cumplimiento"],
            ["Saludo al inicio", f"{protocol_detail.get('greeting_compliance', 0)}%"],
            ["Despedida al final", f"{protocol_detail.get('farewell_compliance', 0)}%"],
            ["Identificación del agente", f"{protocol_detail.get('identification_compliance', 0)}%"],
        ]
        proto_table = Table(proto_data, colWidths=[3.5*inch, 3.5*inch])
        proto_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), ACCENT),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#e0e0e0")),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ]))
        elements.append(proto_table)

    elements.append(Spacer(1, 12))

    # ══════════════════════════════════════
    # 3. DISTRIBUCIÓN DE CALIDAD (gráfico)
    # ══════════════════════════════════════
    class_dist = metrics.get("classification_distribution", {})
    if class_dist:
        elements.append(Paragraph("3. DISTRIBUCIÓN DE CALIDAD", subtitle_style))
        elements.append(HRFlowable(width="100%", thickness=1, color=ACCENT, spaceAfter=8))

        # Tabla + Gráfico de barras
        class_data = [["Clasificación", "Cantidad", "Porcentaje"]]
        class_colors_map = {
            "Excelente": SUCCESS, "Bueno": ACCENT,
            "Regular": WARNING, "Deficiente": DANGER
        }
        for cls_name in ["Excelente", "Bueno", "Regular", "Deficiente"]:
            count = class_dist.get(cls_name, 0)
            pct = round(count / total_audios * 100, 1) if total_audios > 0 else 0
            class_data.append([cls_name, str(count), f"{pct}%"])

        class_table = Table(class_data, colWidths=[2.5*inch, 2*inch, 2.5*inch])
        cls_style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), PRIMARY),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#e0e0e0")),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ]
        for i, cls_name in enumerate(["Excelente", "Bueno", "Regular", "Deficiente"]):
            row = i + 1
            cls_style_cmds.append(('BACKGROUND', (0, row), (0, row), class_colors_map.get(cls_name, LIGHT_BG)))
            cls_style_cmds.append(('TEXTCOLOR', (0, row), (0, row), colors.white))
            cls_style_cmds.append(('FONTNAME', (0, row), (0, row), 'Helvetica-Bold'))

        class_table.setStyle(TableStyle(cls_style_cmds))
        elements.append(class_table)

        # Bar chart
        drawing = Drawing(400, 140)
        chart = VerticalBarChart()
        chart.x = 60
        chart.y = 20
        chart.width = 280
        chart.height = 100
        chart.data = [[
            class_dist.get("Excelente", 0),
            class_dist.get("Bueno", 0),
            class_dist.get("Regular", 0),
            class_dist.get("Deficiente", 0),
        ]]
        chart.categoryAxis.categoryNames = ['Excelente', 'Bueno', 'Regular', 'Deficiente']
        chart.categoryAxis.labels.fontSize = 9
        chart.valueAxis.valueMin = 0
        chart.valueAxis.labels.fontSize = 8
        chart.bars[0].fillColor = ACCENT
        chart.barWidth = 30
        drawing.add(chart)
        elements.append(drawing)
        elements.append(Spacer(1, 12))

    # ══════════════════════════════════════
    # 4. DISTRIBUCIÓN EMOCIONAL
    # ══════════════════════════════════════
    emotion_summary = metrics.get("emotion_summary", {})
    emotion_dist = emotion_summary.get("distribution", {})
    if emotion_dist:
        elements.append(Paragraph("4. DISTRIBUCIÓN EMOCIONAL GENERAL", subtitle_style))
        elements.append(HRFlowable(width="100%", thickness=1, color=ACCENT, spaceAfter=8))

        emo_data = [["Emoción", "Porcentaje"]]
        for emo in ["feliz", "neutral", "triste", "enojado"]:
            pct = emotion_dist.get(emo, 0)
            emo_data.append([emo.capitalize(), f"{pct}%"])

        emo_extra = [
            ["Emoción Predominante", emotion_summary.get("most_common", "N/A").capitalize()],
            ["Alertas Generadas", str(emotion_summary.get("alert_count", 0))],
            ["Tasa de Alertas", f"{emotion_summary.get('alert_rate', 0)}%"],
        ]

        emo_table = Table(emo_data + [["", ""]] + emo_extra, colWidths=[3.5*inch, 3.5*inch])
        emo_style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), PRIMARY),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#e0e0e0")),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ]
        for i, emo in enumerate(["feliz", "neutral", "triste", "enojado"]):
            row = i + 1
            emo_style_cmds.append(('BACKGROUND', (0, row), (0, row), EMOTION_COLORS.get(emo, LIGHT_BG)))
            emo_style_cmds.append(('TEXTCOLOR', (0, row), (0, row), colors.white))

        emo_table.setStyle(TableStyle(emo_style_cmds))
        elements.append(emo_table)

        # Bar chart de emociones
        emo_drawing = Drawing(400, 140)
        emo_chart = VerticalBarChart()
        emo_chart.x = 60
        emo_chart.y = 20
        emo_chart.width = 280
        emo_chart.height = 100
        emo_chart.data = [[
            emotion_dist.get("feliz", 0),
            emotion_dist.get("neutral", 0),
            emotion_dist.get("triste", 0),
            emotion_dist.get("enojado", 0),
        ]]
        emo_chart.categoryAxis.categoryNames = ['Feliz', 'Neutral', 'Triste', 'Enojado']
        emo_chart.categoryAxis.labels.fontSize = 9
        emo_chart.valueAxis.valueMin = 0
        emo_chart.valueAxis.valueMax = 100
        emo_chart.valueAxis.labels.fontSize = 8
        emo_chart.bars[0].fillColor = ACCENT
        emo_chart.barWidth = 30
        emo_drawing.add(emo_chart)
        elements.append(emo_drawing)
        elements.append(Spacer(1, 10))

    elements.append(PageBreak())

    # ══════════════════════════════════════
    # 5. TOP MEJORES Y PEORES AUDIOS
    # ══════════════════════════════════════
    elements.append(Paragraph("5. TOP MEJORES Y PEORES AUDIOS", subtitle_style))
    elements.append(HRFlowable(width="100%", thickness=1, color=ACCENT, spaceAfter=8))

    best = metrics.get("best_audios", [])
    worst = metrics.get("worst_audios", [])

    if best:
        elements.append(Paragraph("Top 5 Mejores:", section_style))
        best_data = [["#", "Archivo", "Score", "Clasificación", "Emoción"]]
        for i, a in enumerate(best[:5]):
            best_data.append([
                str(i+1),
                str(a.get("filename", ""))[:40],
                str(a.get("score", 0)),
                a.get("classification", ""),
                a.get("dominant_emotion", "").capitalize()
            ])
        best_table = Table(best_data, colWidths=[0.4*inch, 3*inch, 0.8*inch, 1.5*inch, 1.3*inch])
        best_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), SUCCESS),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#e0e0e0")),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(best_table)
        elements.append(Spacer(1, 10))

    if worst:
        elements.append(Paragraph("Top 5 Peores:", section_style))
        worst_data = [["#", "Archivo", "Score", "Clasificación", "Emoción"]]
        for i, a in enumerate(worst[:5]):
            worst_data.append([
                str(i+1),
                str(a.get("filename", ""))[:40],
                str(a.get("score", 0)),
                a.get("classification", ""),
                a.get("dominant_emotion", "").capitalize()
            ])
        worst_table = Table(worst_data, colWidths=[0.4*inch, 3*inch, 0.8*inch, 1.5*inch, 1.3*inch])
        worst_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), DANGER),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#e0e0e0")),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(worst_table)

    elements.append(Spacer(1, 12))

    # ══════════════════════════════════════
    # 6. PROBLEMAS COMUNES
    # ══════════════════════════════════════
    top_issues = metrics.get("top_issues", [])
    if top_issues:
        elements.append(Paragraph("6. PROBLEMAS MÁS FRECUENTES", subtitle_style))
        elements.append(HRFlowable(width="100%", thickness=1, color=ACCENT, spaceAfter=8))
        for issue in top_issues:
            elements.append(Paragraph(f"• {_safe_pdf_text(issue)}", body_style))
        elements.append(Spacer(1, 10))

    # ══════════════════════════════════════
    # 7. RECOMENDACIONES GLOBALES
    # ══════════════════════════════════════
    global_recs = metrics.get("global_recommendations", [])
    if global_recs:
        elements.append(Paragraph("7. RECOMENDACIONES GLOBALES", subtitle_style))
        elements.append(HRFlowable(width="100%", thickness=1, color=ACCENT, spaceAfter=8))
        for rec in global_recs:
            elements.append(Paragraph(f"• {_safe_pdf_text(rec)}", body_style))
        elements.append(Spacer(1, 10))

    # ══════════════════════════════════════
    # 8. INTERPRETACIÓN AI (si disponible)
    # ══════════════════════════════════════
    if ai_interpretation:
        elements.append(PageBreak())
        elements.append(Paragraph("8. INTERPRETACIÓN INTELIGENTE (Qwen2.5)", subtitle_style))
        elements.append(HRFlowable(width="100%", thickness=1, color=ACCENT, spaceAfter=8))
        elements.append(Paragraph(
            "<i>Análisis generado por inteligencia artificial:</i>",
            small_style
        ))
        elements.append(Spacer(1, 6))

        # Split interpretation into paragraphs
        for paragraph in ai_interpretation.split("\n"):
            paragraph = _safe_pdf_text(paragraph.strip())
            if paragraph:
                elements.append(Paragraph(paragraph, body_style))
                elements.append(Spacer(1, 4))

        elements.append(Spacer(1, 10))

    # ══════════════════════════════════════
    # 9. DETALLE POR AUDIO
    # ══════════════════════════════════════
    elements.append(PageBreak())
    elements.append(Paragraph("9. DETALLE POR AUDIO ANALIZADO", subtitle_style))
    elements.append(HRFlowable(width="100%", thickness=1, color=ACCENT, spaceAfter=8))

    detail_header = [["#", "Archivo", "Score", "Clasif.", "Emoción", "Alertas"]]
    for i, item in enumerate(history[:100]):  # Máx 100 en el PDF
        qs = item.get("quality_score", {})
        score_val = qs.get("total_score", "N/A") if isinstance(qs, dict) else "N/A"
        cls_val = qs.get("classification", "N/A") if isinstance(qs, dict) else "N/A"
        detail_header.append([
            str(i+1),
            str(item.get("filename", ""))[:35],
            str(score_val),
            cls_val[:10],
            item.get("dominant_emotion", "N/A").capitalize()[:10],
            "Sí" if item.get("has_alert") else "No"
        ])

    detail_table = Table(
        detail_header,
        colWidths=[0.4*inch, 2.8*inch, 0.6*inch, 0.9*inch, 0.9*inch, 0.6*inch],
        repeatRows=1
    )
    detail_style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), PRIMARY),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('LEADING', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor("#e0e0e0")),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]
    for i, item in enumerate(history[:100]):
        row = i + 1
        if row % 2 == 0:
            detail_style_cmds.append(('BACKGROUND', (0, row), (-1, row), LIGHT_BG))
        # Color alert column
        if item.get("has_alert"):
            detail_style_cmds.append(('BACKGROUND', (5, row), (5, row), DANGER))
            detail_style_cmds.append(('TEXTCOLOR', (5, row), (5, row), colors.white))

    detail_table.setStyle(TableStyle(detail_style_cmds))
    elements.append(detail_table)

    # ── FOOTER ──
    elements.append(Spacer(1, 20))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.grey))
    elements.append(Paragraph(
        f"Informe General de Calidad | Voz-a-Texto Emocional API | {date_str}",
        ParagraphStyle('Footer', parent=styles['Normal'], fontSize=7,
                       textColor=colors.grey, alignment=TA_CENTER)
    ))

    doc.build(elements)
    buffer.seek(0)
    return buffer.read()


def _generate_professional_pdf(analysis_data: Dict[str, Any]) -> bytes:
    """Genera un PDF profesional con toda la información del análisis."""
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
        Image, HRFlowable, KeepTogether
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.graphics.shapes import Drawing, Rect, String
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from reportlab.graphics import renderPDF

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=letter,
        leftMargin=0.75*inch, rightMargin=0.75*inch,
        topMargin=0.5*inch, bottomMargin=0.5*inch
    )
    elements = []
    styles = getSampleStyleSheet()

    # Colores corporativos
    PRIMARY = colors.HexColor("#1a237e")
    ACCENT = colors.HexColor("#42a5f5")
    SUCCESS = colors.HexColor("#66bb6a")
    WARNING = colors.HexColor("#ffa726")
    DANGER = colors.HexColor("#ef5350")
    LIGHT_BG = colors.HexColor("#f5f5f5")

    EMOTION_COLORS = {
        "feliz": colors.HexColor("#4caf50"),
        "neutral": colors.HexColor("#42a5f5"),
        "triste": colors.HexColor("#7e57c2"),
        "enojado": colors.HexColor("#ef5350"),
    }

    # Custom styles
    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'],
                                  fontSize=20, textColor=PRIMARY, spaceAfter=6)
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Heading2'],
                                     fontSize=14, textColor=ACCENT, spaceBefore=12, spaceAfter=6)
    body_style = ParagraphStyle('CustomBody', parent=styles['Normal'],
                                 fontSize=10, leading=14)
    small_style = ParagraphStyle('Small', parent=styles['Normal'],
                                  fontSize=8, leading=10, textColor=colors.grey)

    # === LOGO ===
    logo_path = None
    try:
        import glob
        logo_files = glob.glob(os.path.join("data", "company_logo.*"))
        if logo_files:
            logo_path = logo_files[0]
    except:
        pass

    if logo_path and os.path.exists(logo_path):
        try:
            img = Image(logo_path, width=1.5*inch, height=0.75*inch)
            img.hAlign = 'LEFT'
            elements.append(img)
            elements.append(Spacer(1, 6))
        except:
            pass

    # === HEADER ===
    elements.append(Paragraph("Reporte de Análisis Emocional", title_style))
    filename_str = analysis_data.get("filename", analysis_data.get("metadata", {}).get("filename", ""))
    date_str = datetime.now().strftime("%d/%m/%Y %H:%M")
    elements.append(Paragraph(f"Archivo: {filename_str} | Fecha: {date_str}", small_style))
    elements.append(HRFlowable(width="100%", thickness=2, color=PRIMARY, spaceAfter=12))

    # === QUALITY SCORE ===
    quality = analysis_data.get("quality_score", {})
    if quality and isinstance(quality, dict):
        elements.append(Paragraph("Score de Calidad", subtitle_style))
        score = quality.get("total_score", 0)
        classification = quality.get("classification", "N/A")

        score_color = SUCCESS if score >= 70 else WARNING if score >= 50 else DANGER

        score_data = [
            ["Score Total", f"{score}/100", f"Clasificación: {classification}"]
        ]
        breakdown = quality.get("breakdown", {})
        if breakdown:
            for dim_name, dim_data in breakdown.items():
                if isinstance(dim_data, dict):
                    label_map = {
                        "emotional_tone": "Tono Emocional",
                        "keywords": "Palabras Clave",
                        "resolution": "Resolución",
                        "protocol": "Protocolo"
                    }
                    label = label_map.get(dim_name, dim_name)
                    s = dim_data.get("score", 0)
                    m = dim_data.get("max", 0)
                    score_data.append([f"  {label}", f"{s}/{m}", ""])

        score_table = Table(score_data, colWidths=[2.5*inch, 1.5*inch, 3*inch])
        score_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), score_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, -1), LIGHT_BG),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.white),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(score_table)

        # Recommendations
        recs = quality.get("recommendations", [])
        if recs:
            elements.append(Spacer(1, 6))
            for rec in recs[:4]:
                elements.append(Paragraph(f"• {_safe_pdf_text(rec)}", body_style))
        elements.append(Spacer(1, 10))

    # === CALL SUMMARY ===
    summary = analysis_data.get("call_summary", {})
    if summary and isinstance(summary, dict):
        elements.append(Paragraph("Resumen de Llamada", subtitle_style))
        summary_items = [
            ["Motivo", _safe_pdf_text(summary.get("motivo", "N/A"))[:120]],
            ["Desarrollo", _safe_pdf_text(summary.get("desarrollo", "N/A"))[:120]],
            ["Resolucion", _safe_pdf_text(summary.get("resolucion", "N/A"))[:120]],
            ["Satisfaccion", _safe_pdf_text(str(summary.get("satisfaccion_estimada", "N/A")))],
            ["Topico", _safe_pdf_text(summary.get("topico_detectado", "N/A"))],
            ["Duracion", f"{summary.get('duracion_seg', 0):.0f}s"],
        ]
        sum_table = Table(summary_items, colWidths=[1.5*inch, 5.5*inch])
        sum_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), PRIMARY),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.white),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BACKGROUND', (1, 0), (1, -1), LIGHT_BG),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.white),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        elements.append(sum_table)
        elements.append(Spacer(1, 10))

    # === EMOTION DISTRIBUTION CHART ===
    global_emotions = analysis_data.get("global_emotions", {})
    emotion_dist = global_emotions.get("emotion_distribution", {})
    if emotion_dist:
        elements.append(Paragraph("Distribucion Emocional", subtitle_style))

        drawing = Drawing(400, 150)
        chart = VerticalBarChart()
        chart.x = 50
        chart.y = 20
        chart.width = 300
        chart.height = 110
        chart.data = [[
            round(emotion_dist.get("feliz", 0) * 100, 1),
            round(emotion_dist.get("neutral", 0) * 100, 1),
            round(emotion_dist.get("triste", 0) * 100, 1),
            round(emotion_dist.get("enojado", 0) * 100, 1),
        ]]
        chart.categoryAxis.categoryNames = ['Feliz', 'Neutral', 'Triste', 'Enojado']
        chart.categoryAxis.labels.fontSize = 9
        chart.valueAxis.valueMin = 0
        chart.valueAxis.valueMax = 100
        chart.valueAxis.labels.fontSize = 8
        chart.bars[0].fillColor = ACCENT
        chart.barWidth = 30

        drawing.add(chart)
        elements.append(drawing)
        elements.append(Spacer(1, 10))

    # === ALERT ===
    alert = analysis_data.get("alert")
    if alert and isinstance(alert, dict):
        elements.append(Paragraph("Alerta de Escalamiento", subtitle_style))
        sev = alert.get("severity", "media")
        sev_color = DANGER if sev == "alta" else WARNING if sev == "media" else ACCENT
        alert_data_table = [
            ["Severidad", sev.upper()],
        ]
        for reason in alert.get("reasons", [])[:5]:
            alert_data_table.append(["Razón", reason])

        at = Table(alert_data_table, colWidths=[1.5*inch, 5.5*inch])
        at.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), sev_color),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.white),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BACKGROUND', (1, 0), (1, -1), LIGHT_BG),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.white),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(at)
        elements.append(Spacer(1, 10))

    # === SEGMENTS TABLE ===
    segments = analysis_data.get("segments", [])
    if segments:
        elements.append(Paragraph("Transcripcion por Segmentos", subtitle_style))
        seg_header = [["#", "Tiempo", "Hablante", "Emoción", "Texto"]]
        col_widths = [0.3*inch, 0.7*inch, 1*inch, 0.8*inch, 4.2*inch]

        for i, seg in enumerate(segments[:80]):
            text = _safe_pdf_text(seg.get("text_es", seg.get("text", "")))[:80]
            seg_header.append([
                str(i+1),
                f"{seg.get('start', 0):.1f}s",
                seg.get("speaker_label", "")[:12],
                seg.get("emotion", "")[:10],
                text
            ])

        seg_table = Table(seg_header, colWidths=col_widths, repeatRows=1)

        # Build style commands
        style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), PRIMARY),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('LEADING', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor("#e0e0e0")),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]

        # Color-code emotion column per row
        for i, seg in enumerate(segments[:80]):
            row = i + 1
            emotion = seg.get("emotion", "neutral")
            color = EMOTION_COLORS.get(emotion, LIGHT_BG)
            style_cmds.append(('BACKGROUND', (3, row), (3, row), color))
            style_cmds.append(('TEXTCOLOR', (3, row), (3, row), colors.white))
            # Alternate row backgrounds
            if row % 2 == 0:
                style_cmds.append(('BACKGROUND', (0, row), (2, row), LIGHT_BG))
                style_cmds.append(('BACKGROUND', (4, row), (4, row), LIGHT_BG))

        seg_table.setStyle(TableStyle(style_cmds))
        elements.append(seg_table)

    # === FOOTER ===
    elements.append(Spacer(1, 20))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.grey))
    elements.append(Paragraph(
        f"Generado por Voz-a-Texto Emocional API | {date_str}",
        ParagraphStyle('Footer', parent=styles['Normal'], fontSize=7,
                       textColor=colors.grey, alignment=TA_CENTER)
    ))

    doc.build(elements)
    buffer.seek(0)
    return buffer.read()


@router.post("/export/excel", tags=["export"])
async def export_excel_endpoint(
    data: str = Form(...),
    filename: str = Form("reporte_analisis")
) -> Response:
    """
    Exporta reporte Excel profesional con múltiples hojas y gráfico.
    El campo 'data' debe ser JSON con la respuesta completa del análisis.
    """
    try:
        analysis_data = json.loads(data)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="JSON inválido")

    try:
        excel_bytes = _generate_excel_report(analysis_data)
        return Response(
            content=excel_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"}
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl no instalado. Ejecutar: pip install openpyxl")
    except Exception as e:
        logger.error(f"Error generando Excel: {e}")
        raise HTTPException(status_code=500, detail=str(e))


def _generate_excel_report(analysis_data: Dict[str, Any]) -> bytes:
    """Genera un reporte Excel profesional con múltiples hojas."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Colors
    PRIMARY_FILL = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
    ACCENT_FILL = PatternFill(start_color="42a5f5", end_color="42a5f5", fill_type="solid")
    SUCCESS_FILL = PatternFill(start_color="66bb6a", end_color="66bb6a", fill_type="solid")
    WARNING_FILL = PatternFill(start_color="ffa726", end_color="ffa726", fill_type="solid")
    DANGER_FILL = PatternFill(start_color="ef5350", end_color="ef5350", fill_type="solid")
    LIGHT_FILL = PatternFill(start_color="f5f5f5", end_color="f5f5f5", fill_type="solid")
    WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
    HEADER_FONT = Font(bold=True, size=11)
    TITLE_FONT = Font(bold=True, size=14, color="1a237e")
    thin_border = Border(
        left=Side(style='thin', color='e0e0e0'),
        right=Side(style='thin', color='e0e0e0'),
        top=Side(style='thin', color='e0e0e0'),
        bottom=Side(style='thin', color='e0e0e0'),
    )

    EMOTION_FILLS = {
        "feliz": PatternFill(start_color="c8e6c9", end_color="c8e6c9", fill_type="solid"),
        "neutral": PatternFill(start_color="bbdefb", end_color="bbdefb", fill_type="solid"),
        "triste": PatternFill(start_color="d1c4e9", end_color="d1c4e9", fill_type="solid"),
        "enojado": PatternFill(start_color="ffcdd2", end_color="ffcdd2", fill_type="solid"),
    }

    # === HOJA 1: RESUMEN ===
    ws1 = wb.active
    ws1.title = "Resumen"
    ws1.sheet_properties.tabColor = "1a237e"

    ws1['A1'] = "Reporte de Análisis Emocional"
    ws1['A1'].font = TITLE_FONT
    ws1.merge_cells('A1:D1')

    ws1['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws1['A2'].font = Font(size=9, color="888888")

    # Quality Score section
    row = 4
    ws1.cell(row=row, column=1, value="Score de Calidad").font = Font(bold=True, size=12, color="1a237e")
    quality = analysis_data.get("quality_score", {})
    if isinstance(quality, dict):
        row += 1
        for label, key in [("Score Total", "total_score"), ("Clasificación", "classification")]:
            ws1.cell(row=row, column=1, value=label).font = HEADER_FONT
            ws1.cell(row=row, column=1).fill = PRIMARY_FILL
            ws1.cell(row=row, column=1).font = WHITE_FONT
            val = quality.get(key, "N/A")
            ws1.cell(row=row, column=2, value=f"{val}/100" if key == "total_score" else val)
            score_val = quality.get("total_score", 0)
            if key == "total_score":
                fill = SUCCESS_FILL if score_val >= 70 else WARNING_FILL if score_val >= 50 else DANGER_FILL
                ws1.cell(row=row, column=2).fill = fill
                ws1.cell(row=row, column=2).font = Font(bold=True, color="FFFFFF", size=14)
            row += 1

        # Breakdown
        breakdown = quality.get("breakdown", {})
        label_map = {
            "emotional_tone": "Tono Emocional",
            "keywords": "Palabras Clave",
            "resolution": "Resolución",
            "protocol": "Protocolo"
        }
        for dim_name, dim_data in breakdown.items():
            if isinstance(dim_data, dict):
                ws1.cell(row=row, column=1, value=f"  {label_map.get(dim_name, dim_name)}")
                ws1.cell(row=row, column=2, value=f"{dim_data.get('score', 0)}/{dim_data.get('max', 0)}")
                row += 1

        # Recommendations
        recs = quality.get("recommendations", [])
        if recs:
            row += 1
            ws1.cell(row=row, column=1, value="Recomendaciones").font = Font(bold=True, size=11, color="1a237e")
            row += 1
            for rec in recs:
                ws1.cell(row=row, column=1, value=rec)
                ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
                row += 1

    # Call Summary section
    summary = analysis_data.get("call_summary", {})
    if isinstance(summary, dict):
        row += 1
        ws1.cell(row=row, column=1, value="Resumen de Llamada").font = Font(bold=True, size=12, color="1a237e")
        row += 1
        for label, key in [("Motivo", "motivo"), ("Desarrollo", "desarrollo"),
                           ("Resolución", "resolucion"), ("Satisfacción", "satisfaccion_estimada"),
                           ("Tópico", "topico_detectado")]:
            ws1.cell(row=row, column=1, value=label).font = HEADER_FONT
            ws1.cell(row=row, column=1).fill = ACCENT_FILL
            ws1.cell(row=row, column=1).font = WHITE_FONT
            val = str(summary.get(key, "N/A"))[:100]
            ws1.cell(row=row, column=2, value=val)
            ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            row += 1

    # Emotion Distribution with chart
    global_emotions = analysis_data.get("global_emotions", {})
    emotion_dist = global_emotions.get("emotion_distribution", {})
    if emotion_dist:
        row += 1
        ws1.cell(row=row, column=1, value="Distribución Emocional").font = Font(bold=True, size=12, color="1a237e")
        row += 1
        chart_start = row
        emotions = ["feliz", "neutral", "triste", "enojado"]
        ws1.cell(row=row, column=1, value="Emoción").font = WHITE_FONT
        ws1.cell(row=row, column=1).fill = PRIMARY_FILL
        ws1.cell(row=row, column=2, value="Porcentaje").font = WHITE_FONT
        ws1.cell(row=row, column=2).fill = PRIMARY_FILL
        row += 1
        for emo in emotions:
            ws1.cell(row=row, column=1, value=emo.capitalize())
            ws1.cell(row=row, column=1).fill = EMOTION_FILLS.get(emo, LIGHT_FILL)
            pct = round(emotion_dist.get(emo, 0) * 100, 1)
            ws1.cell(row=row, column=2, value=pct)
            row += 1

        # Chart
        chart = BarChart()
        chart.style = 10
        chart.title = "Distribución Emocional"
        chart.y_axis.title = "Porcentaje (%)"
        chart.y_axis.scaling.max = 100
        data_ref = Reference(ws1, min_col=2, min_row=chart_start, max_row=chart_start + 4)
        cats_ref = Reference(ws1, min_col=1, min_row=chart_start + 1, max_row=chart_start + 4)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        chart.width = 14
        chart.height = 10
        ws1.add_chart(chart, f"D{chart_start}")

    # Column widths
    ws1.column_dimensions['A'].width = 22
    ws1.column_dimensions['B'].width = 25
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 20

    # === HOJA 2: SEGMENTOS ===
    ws2 = wb.create_sheet("Segmentos")
    ws2.sheet_properties.tabColor = "42a5f5"

    headers = ["#", "Tiempo", "Hablante", "Emoción", "Score", "Texto"]
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = PRIMARY_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    segments = analysis_data.get("segments", [])
    for i, seg in enumerate(segments):
        row = i + 2
        ws2.cell(row=row, column=1, value=i+1).border = thin_border
        ws2.cell(row=row, column=2, value=f"{seg.get('start', 0):.1f}s").border = thin_border
        ws2.cell(row=row, column=3, value=seg.get("speaker_label", "")).border = thin_border

        emotion = seg.get("emotion", "neutral")
        emo_cell = ws2.cell(row=row, column=4, value=emotion)
        emo_cell.fill = EMOTION_FILLS.get(emotion, LIGHT_FILL)
        emo_cell.border = thin_border

        score_val = seg.get("emotions", {}).get("fused", {}).get("score", 0)
        ws2.cell(row=row, column=5, value=round(score_val, 3) if score_val else 0).border = thin_border
        ws2.cell(row=row, column=6, value=seg.get("text_es", seg.get("text", ""))).border = thin_border

        if row % 2 == 0:
            for c in [1, 2, 3, 5, 6]:
                ws2.cell(row=row, column=c).fill = LIGHT_FILL

    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 15
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 10
    ws2.column_dimensions['F'].width = 60
    ws2.auto_filter.ref = f"A1:F{len(segments)+1}"

    # === HOJA 3: ALERTAS ===
    alert = analysis_data.get("alert")
    if alert and isinstance(alert, dict):
        ws3 = wb.create_sheet("Alertas")
        ws3.sheet_properties.tabColor = "ef5350"
        ws3['A1'] = "🚨 Alerta de Escalamiento"
        ws3['A1'].font = Font(bold=True, size=14, color="ef5350")
        ws3.merge_cells('A1:C1')

        ws3.cell(row=3, column=1, value="Severidad").font = WHITE_FONT
        ws3.cell(row=3, column=1).fill = DANGER_FILL
        ws3.cell(row=3, column=2, value=alert.get("severity", "").upper())

        row = 4
        for reason in alert.get("reasons", []):
            ws3.cell(row=row, column=1, value="Razón").font = HEADER_FONT
            ws3.cell(row=row, column=2, value=reason)
            ws3.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            row += 1

        ws3.column_dimensions['A'].width = 15
        ws3.column_dimensions['B'].width = 40
        ws3.column_dimensions['C'].width = 20

    # Save
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


@router.post("/export/json", tags=["export"])
async def export_json(
    segments: str = Form(...),
    global_emotions: str = Form("{}"),
    speaker_stats: str = Form("{}"),
    metadata: str = Form("{}"),
    filename: str = Form("analisis_emocional")
) -> Response:
    """
    Exporta los segmentos de la transcripción en formato JSON.
    """

    try:
        data= ExportData(
            segments=json.loads(segments),
            global_emotions=json.loads(global_emotions),
            speaker_stats=json.loads(speaker_stats),
            metadata=json.loads(metadata),
            filename=filename
        )
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Formato de segmentos inválido")

    exporter = get_export_manager()
    content = exporter.export_json(data)

    return Response(content=content, media_type="application/json",headers={"Content-Disposition": f"attachment; filename={filename}.json"})

#alamacenamiento termporal de las sesiones 

@router.post("/store_session", tags=["sessions"])
async def store_session(namer:str = None, data:str = None) -> Dict[str,Any]:

    """
    Almacena una sesión en el almacenamiento temporal.
    """
    session_id = str(uuid.uuid4())
    time_stamp = datetime.now()

    try:
        session_data=json.loads(data)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Formato de datos inválido")
    

    session = {
        "id": session_id,
        "name": namer or f"Sesión {time_stamp.strftime('%Y-%m-%d %H:%M')}",
        "created_at": time_stamp.isoformat(),
        "updated_at": time_stamp.isoformat(),
        "data": session_data
    }

    _sessions_store[session_id] = session
    return {"status": "success",
        "session_id": session_id,
        "name": session["name"],
        "created_at": session["created_at"]}


@router.get("/sessions/{session_id}", tags=["Sessions"])
async def get_session(session_id: str) -> Dict[str, Any]:
    """
    Obtiene una sesión por ID.
    """
    if session_id not in _sessions_store:
        raise HTTPException(status_code=404, detail="Sesión no encontrada")
    
    return _sessions_store[session_id]



@router.put("/sessions/{session_id}", tags=["Sessions"])
async def update_session(
    session_id: str,
    name: str = Form(None),
    data: str = Form(None)
) -> Dict[str, Any]:
    """
    Actualiza una sesion existente.
    """
    if session_id not in _sessions_store:
        raise HTTPException(status_code=404, detail="Sesión no encontrada")
    
    session = _sessions_store[session_id]
    
    if name:
        session["name"] = name
    
    if data:
        try:
            session["data"] = json.loads(data)
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail="JSON inválido")
    
    session["updated_at"] = datetime.now().isoformat()
    
    return {
        "status": "success",
        "session_id": session_id,
        "updated_at": session["updated_at"]
    }

@router.delete("/sessions/{session_id}", tags=["Sessions"])
async def delete_session(session_id: str) -> Dict[str, Any]:
    """
    Elimina una sesion existente.
    """
    if session_id not in _sessions_store:
        raise HTTPException(status_code=404, detail="Sesión no encontrada")
    
    del _sessions_store[session_id]
    
    return {"status": "success", "message": "Sesión eliminada exitosamente"}

@router.get("/sessions", tags=["Sessions"])
async def list_sessions(
    limit: int = Query(20, ge=1, le=100)
) -> Dict[str, Any]:
    """
    Lista todas las sesiones.
    """
    sessions = sorted(
        _sessions_store.values(),
        key=lambda x: x["updated_at"],
        reverse=True
    )[:limit]
    
    return {
        "sessions": [
            {
                "id": s["id"],
                "name": s["name"],
                "created_at": s["created_at"],
                "updated_at": s["updated_at"],
                "segments_count": len(s["data"].get("segments", []))
            }
            for s in sessions
        ],
        "total": len(_sessions_store)
    }

@router.post("/transcription/update-segment", tags=["Transcription"])
async def update_segment(
    session_id: str = Form(...),
    segment_index: int = Form(...),
    text: str = Form(None),
    speaker_label: str = Form(None),
    speaker_id: int = Form(None)
) -> Dict[str, Any]:
    """
    Actualiza un segmento de transcripcin.
    Permite corregir texto o reasignar hablante.
    """
    if session_id not in _sessions_store:
        raise HTTPException(status_code=404, detail="Sesión no encontrada")
    
    session = _sessions_store[session_id]
    segments = session["data"].get("segments", [])
    
    if segment_index < 0 or segment_index >= len(segments):
        raise HTTPException(status_code=400, detail="Índice de segmento inválido")
    
    segment = segments[segment_index]
    
    if text is not None:
        segment["text_es"] = text
        segment["text"] = text
        segment["_edited"] = True
    
    if speaker_label is not None:
        segment["speaker_label"] = speaker_label
        segment["_speaker_edited"] = True
    
    if speaker_id is not None:
        segment["speaker_id"] = speaker_id
    
    session["updated_at"] = datetime.now().isoformat()
    
    return {
        "status": "success",
        "segment_index": segment_index,
        "updated_segment": segment
    }


@router.post("/transcription/merge-speakers", tags=["Transcription"])
async def merge_speakers(
    session_id: str = Form(...),
    source_speaker_id: int = Form(...),
    target_speaker_id: int = Form(...),
    target_label: str = Form(None)
) -> Dict[str, Any]:
    """
    Fusiona dos hablantes en uno.

    """
    if session_id not in _sessions_store:
        raise HTTPException(status_code=404, detail="Sesión no encontrada")
    
    session = _sessions_store[session_id]
    segments = session["data"].get("segments", [])
    
    merged_count = 0
    for segment in segments:
        if segment.get("speaker_id") == source_speaker_id:
            segment["speaker_id"] = target_speaker_id
            if target_label:
                segment["speaker_label"] = target_label
            merged_count += 1
    
    session["updated_at"] = datetime.now().isoformat()
    
    return {
        "status": "success",
        "merged_segments": merged_count,
        "source_speaker_id": source_speaker_id,
        "target_speaker_id": target_speaker_id
    }


#End point para la transcripcion
@router.post("/transcribe/with-provider", tags=["Transcription"])
async def transcribe_with_provider(
    file: UploadFile = File(...),
    provider: str = Form("local"),
    api_key: str = Form(None),
    language: str = Form("es")
) -> Dict[str, Any]:
    """
    Transcribe audio usando el proveedor especificado.
    """
    import tempfile
    import aiofiles
    
    provider_map = {
        "local": TranscriptionProvider.LOCAL,
        "openai": TranscriptionProvider.OPENAI,
        "groq": TranscriptionProvider.GROQ
    }
    
    if provider not in provider_map:
        raise HTTPException(status_code=400, detail="Proveedor no valido")
    
    # Verificar API key para proveedores cloud
    if provider in ["openai", "groq"] and not api_key:
        service = get_transcription_service()
        if not service.has_api_key(provider_map[provider]):
            raise HTTPException(
                status_code=400,
                detail=f"Se requiere API key para {provider}"
            )
    
    # Guardar archivo temporal
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(file.filename)[1]) as tmp:
            content = await file.read()
            tmp.write(content)
            tmp_path = tmp.name
        
        # Transcribir
        service = get_transcription_service()
        result = await service.transcribe(
            audio_path=tmp_path,
            provider=provider_map[provider],
            language=language,
            api_key=api_key
        )
        
        return {
            "status": "success",
            "provider": result.provider,
            "text": result.text,
            "segments": result.segments,
            "language": result.language,
            "duration": result.duration,
            "processing_time": result.processing_time,
            "cost_estimate": result.cost_estimate
        }
        
    except Exception as e:
        logger.error(f"Error en transcripción: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.remove(tmp_path)
